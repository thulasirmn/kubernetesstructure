from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "AKS Sizing & Cost"

# palette
PRIMARY = "1F4E78"; ACCENT = "2E75B6"; LIGHT = "DDEBF7"
INPUTBG = "FFF2CC"; INPUTTX = "7F6000"
GREENBG = "C6EFCE"; GREENTX = "006100"
GREY = "808080"; ZEBRA = "F2F7FC"

title_f = Font(bold=True, size=15, color="FFFFFF")
hdr_f = Font(bold=True, size=12, color="FFFFFF")
input_f = Font(bold=True, color=INPUTTX)
calc_f = Font(color="333333")
res_f = Font(bold=True, size=12, color=GREENTX)
note_f = Font(italic=True, size=9, color=GREY)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

widths = {"A": 3, "B": 38, "C": 15, "D": 14, "E": 15, "F": 16, "G": 13, "H": 14, "I": 13, "J": 12, "K": 14}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

R = Alignment(horizontal="right")
CEN = Alignment(horizontal="center")
USD0 = '"$"#,##0'
USD2 = '"$"#,##0.00'


def title(r, text):
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]; c.value = text; c.font = title_f
    c.fill = PatternFill("solid", fgColor=PRIMARY); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22


def band(r, text):
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]; c.value = text; c.font = hdr_f; c.fill = PatternFill("solid", fgColor=ACCENT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 18


def lbl(r, text, col="B", font=calc_f):
    c = ws[f"{col}{r}"]; c.value = text; c.font = font; return c


def inp(r, val, fmt="#,##0.00", col="C"):
    c = ws[f"{col}{r}"]; c.value = val; c.fill = PatternFill("solid", fgColor=INPUTBG)
    c.font = input_f; c.number_format = fmt; c.border = border; c.alignment = R; return c


def calc(r, formula, fmt="#,##0.00", col="C", font=None, fill=None):
    c = ws[f"{col}{r}"]; c.value = formula; c.number_format = fmt; c.font = font or calc_f; c.alignment = R
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    return c


def note(r, text):
    ws.merge_cells(f"B{r}:K{r}")
    c = ws[f"B{r}"]; c.value = text; c.font = note_f; return c


# ===== TITLE =====
title(1, "SRW on AKS — Cluster Sizing & Cost vs TRE App Service")
note(2, "Yellow = inputs (edit freely). Green = pick the node SKU from the dropdown. All capacity, node counts and costs recalculate.")

# ===== CAPACITY INPUTS =====
band(4, "INPUTS — CAPACITY")
lbl(5, "Workspaces"); inp(5, 500, "#,##0")
lbl(6, "Users per workspace"); inp(6, 10, "#,##0")
lbl(7, "Total users"); calc(7, "=C5*C6", "#,##0")
lbl(8, "Peak concurrency (% of users active)"); inp(8, 0.40, "0%")
lbl(9, "Avg apps open per active user"); inp(9, 1.0, "#,##0.0")
lbl(10, "Concurrent sessions (peak)"); calc(10, "=ROUND(C7*C8*C9,0)", "#,##0", font=Font(bold=True))
lbl(11, "Per-pod CPU request (vCPU)"); inp(11, 0.5, "#,##0.00")
lbl(12, "Per-pod RAM request (GiB)"); inp(12, 1, "#,##0.0")
lbl(13, "Per-pod CPU limit (vCPU, burst)"); inp(13, 2, "#,##0.0")
lbl(14, "Per-pod RAM limit (GiB, burst)"); inp(14, 4, "#,##0.0")
lbl(15, "Node headroom (burst + fragmentation)"); inp(15, 0.35, "0%")
lbl(16, "Node CPU allocatable factor"); inp(16, 0.95, "0%")
lbl(17, "Node RAM allocatable factor"); inp(17, 0.88, "0%")
lbl(18, "DaemonSet vCPU per node"); inp(18, 0.4, "#,##0.00")
lbl(19, "DaemonSet RAM per node (GiB)"); inp(19, 0.5, "#,##0.0")

# ===== COST / UTILIZATION INPUTS =====
band(21, "INPUTS — COST & UTILIZATION")
lbl(22, "Peak business hours / month"); inp(22, 220, "#,##0")
lbl(23, "Off-peak hours / month"); calc(23, "=730-C22", "#,##0")
lbl(24, "Off-peak baseline concurrency (%)"); inp(24, 0.10, "0%")
lbl(25, "Off-peak concurrent sessions"); calc(25, "=ROUND(C7*C24*C9,0)", "#,##0")
lbl(26, "AKS fixed platform cost ($/mo)"); inp(26, 520, USD0)
lbl(27, "TRE App Service Plan ($/mo, P2v3)"); inp(27, 267.18, USD2)
lbl(28, "TRE App Service Plans per workspace"); inp(28, 1, "#,##0")

# ===== AGGREGATE =====
band(30, "AGGREGATE CAPACITY AT PEAK")
lbl(31, "Reserved vCPU (sessions x request)"); calc(31, "=C10*C11", "#,##0", font=Font(bold=True))
lbl(32, "Reserved RAM (GiB)"); calc(32, "=C10*C12", "#,##0", font=Font(bold=True))
lbl(33, "Reserved RAM (TB)"); calc(33, "=C32/1024", "#,##0.0")
lbl(34, "Burst-max vCPU (every pod at its limit)"); calc(34, "=C10*C13", "#,##0")

# ===== NODE SKU TABLE =====
band(36, "NODE SKU COMPARISON  (capacity + monthly cost, auto-computed)")
heads = ["Node SKU", "vCPU", "RAM (GiB)", "Usable vCPU", "Usable RAM (GiB)",
         "Sessions/node", "Peak nodes", "Off-peak nodes", "Node $/hr", "AKS $/mo"]
for j, h in enumerate(heads):
    c = ws.cell(row=37, column=2 + j, value=h); c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=ACCENT); c.border = border
    c.alignment = CEN if j else Alignment(horizontal="left")

# SKU: name, vCPU, RAM GiB, PAYG $/hr (Linux, East US, approx)
skus = [
    ("D8s_v3", 8, 32, 0.384),
    ("D16s_v3", 16, 64, 0.768),
    ("D32s_v3", 32, 128, 1.536),
    ("F16s_v2", 16, 32, 0.677),
    ("F32s_v2", 32, 64, 1.354),
]
first, last = 38, 38 + len(skus) - 1
for i, (name, vcpu, ram, price) in enumerate(skus):
    r = 38 + i
    fill = ZEBRA if i % 2 else None
    cB = ws[f"B{r}"]; cB.value = name; cB.border = border
    if fill: cB.fill = PatternFill("solid", fgColor=fill)
    calc(r, vcpu, "#,##0", col="C")
    calc(r, ram, "#,##0", col="D")
    calc(r, f"=C{r}*$C$16-$C$18", "#,##0.0", col="E")
    calc(r, f"=D{r}*$C$17-$C$19", "#,##0.0", col="F")
    calc(r, f"=MIN(ROUNDDOWN(E{r}/$C$11,0),ROUNDDOWN(F{r}/$C$12,0))", "#,##0", col="G")
    calc(r, f"=ROUNDUP($C$10*(1+$C$15)/G{r},0)", "#,##0", col="H")
    calc(r, f"=ROUNDUP($C$25*(1+$C$15)/G{r},0)", "#,##0", col="I")
    # node price input cell
    cJ = ws[f"J{r}"]; cJ.value = price; cJ.fill = PatternFill("solid", fgColor=INPUTBG)
    cJ.font = input_f; cJ.number_format = USD2; cJ.border = border; cJ.alignment = R
    # AKS $/mo for this SKU
    calc(r, f"=(H{r}*$C$22+I{r}*$C$23)*J{r}+$C$26", USD0, col="K")
    for col in "CDEFGHIK":
        ws[f"{col}{r}"].border = border
        if fill:
            ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=fill)

# ===== SELECTED SKU =====
band(44, "RECOMMENDED CLUSTER  (driven by the SKU you pick below)")
lbl(45, "Selected node SKU")
sel = ws["C45"]; sel.value = "D16s_v3"; sel.fill = PatternFill("solid", fgColor=GREENBG)
sel.font = Font(bold=True, color=GREENTX); sel.alignment = CEN; sel.border = border
dv = DataValidation(type="list", formula1='"D8s_v3,D16s_v3,D32s_v3,F16s_v2,F32s_v2"', allow_blank=False)
ws.add_data_validation(dv); dv.add(sel)

MR = f"MATCH(C45,B{first}:B{last},0)"
lbl(46, "Sessions per node"); calc(46, f"=INDEX(G{first}:G{last},{MR})", "#,##0")
lbl(47, "User node pool — peak nodes"); calc(47, f"=INDEX(H{first}:H{last},{MR})", "#,##0", font=res_f, fill=GREENBG)
lbl(48, "User node pool — off-peak nodes"); calc(48, f"=INDEX(I{first}:I{last},{MR})", "#,##0")
lbl(49, "Selected node $/hr"); calc(49, f"=INDEX(J{first}:J{last},{MR})", USD2)
lbl(50, "Peak vCPU provisioned (user pool)"); calc(50, f"=C47*INDEX(C{first}:C{last},{MR})", "#,##0")
lbl(51, "Peak RAM provisioned (GiB, user pool)"); calc(51, f"=C47*INDEX(D{first}:D{last},{MR})", "#,##0")
lbl(52, "System node pool (fixed)"); c = ws["C52"]; c.value = "3 x D4s_v3"; c.alignment = R; c.font = calc_f

# ===== COST COMPARISON =====
band(54, "COST COMPARISON — SRW on AKS  vs  TRE App Service")
for j, h in enumerate(["", "Per month", "Per year"]):
    c = ws.cell(row=55, column=2 + j, value=h); c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=ACCENT); c.border = border; c.alignment = CEN if j else Alignment(horizontal="left")
lbl(56, "TRE — one App Service Plan per workspace (24x7)")
calc(56, "=C5*C28*C27", USD0); calc(56, "=C56*12", USD0, col="D")
lbl(57, "SRW on AKS — shared cluster (selected SKU)")
calc(57, f"=INDEX(K{first}:K{last},{MR})", USD0); calc(57, "=C57*12", USD0, col="D")
lbl(58, "Savings", font=Font(bold=True, color=GREENTX))
calc(58, "=C56-C57", USD0, font=res_f, fill=GREENBG); calc(58, "=D56-D57", USD0, col="D", font=res_f, fill=GREENBG)
lbl(59, "% reduction", font=Font(bold=True, color=GREENTX))
calc(59, "=(C56-C57)/C56", "0%", font=res_f, fill=GREENBG)
for col in ("B", "C", "D"):
    for r in (56, 57, 58, 59):
        ws[f"{col}{r}"].border = border

# ===== SENSITIVITY =====
band(61, "CONCURRENCY SENSITIVITY  (selected SKU; off-peak baseline held fixed)")
for j, h in enumerate(["Peak concurrency", "Concurrent sessions", "Peak nodes", "AKS $/mo", "Savings $/mo vs TRE"]):
    c = ws.cell(row=62, column=2 + j, value=h); c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor=LIGHT); c.border = border; c.alignment = CEN if j else Alignment(horizontal="left")
for i, conc in enumerate([0.10, 0.20, 0.40, 0.60, 1.00]):
    r = 63 + i
    ws[f"B{r}"].value = conc; ws[f"B{r}"].number_format = "0%"
    calc(r, f"=ROUND($C$7*{conc}*$C$9,0)", "#,##0", col="C")
    calc(r, f"=ROUNDUP(C{r}*(1+$C$15)/$C$46,0)", "#,##0", col="D")
    calc(r, f"=(D{r}*$C$22+$C$48*$C$23)*$C$49+$C$26", USD0, col="E")
    calc(r, f"=$C$56-E{r}", USD0, col="F")
    for col in "BCDEF":
        ws[f"{col}{r}"].border = border

note(69, "Costs: PAYG list rates (Linux, East US) — D8s_v3 $0.384/hr; node $/hr cells are editable (swap in your EA/CSP rate). "
         "AKS $/mo = (peak nodes x peak hrs + off-peak nodes x off-peak hrs) x node $/hr + fixed. Node counts are the autoscaler "
         "PEAK ceiling, not 24/7 spend. Reserved Instances on the baseline pool cut compute a further ~40%. Reminder: 500 "
         "workspaces = 500 storage accounts > Azure default 250/region (request a quota increase).")

out = r"D:\kubernetesstructure\SecureResearchWorkspace\SecureResearchWorkspace\SRW_AKS_Sizing.xlsx"
wb.save(out)
print("saved", out)
